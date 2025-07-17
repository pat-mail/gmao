import tkinter as tk
from tkinter import simpledialog, messagebox
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image as XLImage
import os

entries = []  # globale

def export_by_palette(conn):
    cursor = conn.cursor()

    palette_input = simpledialog.askstring("Palette", "Entrez le numéro de palette :")
    if not palette_input:
        return

    palette_clean = palette_input.strip().rstrip(".0")

    cursor.execute("""
        SELECT n_venus, lot, isotope, modele, date_livraison, decroissance, date_depart
        FROM genes_uses
        WHERE TRIM(REPLACE(palette, '.0', '')) = ?
        ORDER BY lot
    """, (palette_clean,))
    rows = cursor.fetchall()

    if not rows:
        messagebox.showinfo("Aucun résultat", "Aucun générateur trouvé pour cette palette.")
        return

    wb = Workbook()
    ws = wb.active

    try:
        date_obj = datetime.strptime(palette_clean, "%d%m%y")
        date_str = date_obj.strftime("%d-%m-%y")
    except ValueError:
        date_str = palette_clean

    titre = f"Palette du {date_str}"
    ws.merge_cells('A1:G1')
    cell_title = ws['A1']
    cell_title.value = titre
    cell_title.font = Font(size=14, bold=True)
    cell_title.alignment = Alignment(horizontal="center")

    logo_path = "logo.png"
    if os.path.exists(logo_path):
        logo = XLImage(logo_path)
        logo.anchor = 'H1'
        ws.add_image(logo)

    headers = ["N° Vénus", "Lot", "Isotope", "Modèle", "Date livraison", "Décroissance", "Départ"]
    ws.append(headers)
    for cell in ws[2]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append(row)

    from openpyxl.utils import get_column_letter

    for i, column_cells in enumerate(ws.iter_cols(min_row=2), start=1):
        max_length = 0
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = max(12, max_length + 2)

    nom_fichier = f"palette_{palette_clean}.xlsx"
    try:
        wb.save(nom_fichier)
        messagebox.showinfo("Export réussi", f"Fichier '{nom_fichier}' créé avec succès.")
    except Exception as e:
        messagebox.showerror("Erreur d'export", str(e))


def load(frame_left, frame_right, conn):
    for w in frame_left.winfo_children():
        w.destroy()
    for w in frame_right.winfo_children():
        w.destroy()

    cursor = conn.cursor()
    cursor.execute("""SELECT n_venus, lot, isotope, modele, date_livraison, decroissance, palette, date_depart FROM genes_uses""")
    rows = cursor.fetchall()

    container_left = tk.Frame(frame_left)
    container_left.pack(fill='both', expand=True)

    scrollbar = tk.Scrollbar(container_left)
    scrollbar.pack(side='right', fill='y')

    listbox = tk.Listbox(container_left, yscrollcommand=scrollbar.set, font=("Courier", 8))
    listbox.pack(side='left', fill='both', expand=True)
    scrollbar.config(command=listbox.yview)

    ctrl_left = tk.Frame(container_left)
    ctrl_left.pack(fill='x', pady=5)
    tk.Button(ctrl_left, text="Rafraîchir", command=lambda: load(frame_left, frame_right, conn)).pack(side='right', padx=2)

    data_by_index = {}
    today = datetime.today().date()

    for i, row in enumerate(rows):
        n_venus, lot, isotope, modele, date_livraison, decroissance_str, palette, date_depart_str = row
        try:
            decroissance = datetime.strptime(decroissance_str, "%Y-%m-%d %H:%M:%S").date() if decroissance_str else None
        except:
            decroissance = None
        try:
            date_depart = datetime.strptime(date_depart_str, "%Y-%m-%d %H:%M:%S").date() if date_depart_str else None
        except:
            date_depart = None

        froid_le = decroissance + timedelta(days=90) if decroissance else None
        palette_str = palette[:-2] if palette and isinstance(palette, str) and palette.endswith('.0') else str(palette)
        text = f"{n_venus:<8} | {lot:<10} | Modèle: {modele:<10} | Palette: {palette_str:<5} | Départ: {str(date_depart) if date_depart else 'N/A'}"
        listbox.insert(tk.END, text)
        data_by_index[i] = (row, froid_le, date_depart)

        # Définir la couleur
        if date_depart:
            listbox.itemconfig(i, fg="black")  # Expédié
        elif froid_le:
            if today < froid_le:
                listbox.itemconfig(i, fg="red")  # Encore chaud
            else:
                listbox.itemconfig(i, fg="green")  # Refroidi
        else:
            listbox.itemconfig(i, fg="orange")  # Pas de date de décroissance




    details_frame = tk.Frame(frame_right)
    details_frame.pack(fill='both', expand=True)

    def refresh():
        load(frame_left, frame_right, conn)

    def display_details(row, froid_le):
        for w in details_frame.winfo_children():
            w.destroy()
        champs = ["n_venus", "lot", "isotope", "modele", "date_livraison", "decroissance", "palette", "date_depart", "froid_le"]
        valeurs = list(row) + [froid_le.strftime("%Y-%m-%d") if froid_le else "N/A"]
        for i, (champ, val) in enumerate(zip(champs, valeurs)):
            tk.Label(details_frame, text=f"{champ}:", font=("Arial", 10, "bold")).grid(row=i, column=0, sticky='w', padx=5, pady=2)
            tk.Label(details_frame, text=str(val)).grid(row=i, column=1, sticky='w', padx=5, pady=2)

        btn_frame = tk.Frame(details_frame)
        btn_frame.grid(row=len(champs), column=0, columnspan=2, pady=10)

        form_frame = tk.Frame(details_frame)  # <-- Ajout ici
        form_frame.grid(row=len(champs)+1, column=0, columnspan=2, pady=10)

        tk.Button(btn_frame, text="Ajouter", command=lambda: show_form(None, form_frame)).grid(row=0, column=0, padx=5)
        tk.Button(btn_frame, text="Modifier", command=lambda: show_form(row, form_frame)).grid(row=0, column=1)
        tk.Button(btn_frame, text="Supprimer", command=lambda: delete_selected()).grid(row=0, column=2, padx=5)

        tk.Button(btn_frame, text="Palette", command=lambda: export_by_palette(conn)).grid(row=1, column=0, columnspan=3, pady=(10, 0))

    def show_details(event):
        sel = listbox.curselection()
        if not sel:
            return
        idx = sel[0]
        row, froid_le, date_depart = data_by_index[idx]
        display_details(row, froid_le)

    def show_form(data=None, form_frame=None):
        for widget in form_frame.winfo_children():
            widget.destroy()

        labels = ["n_venus", "lot", "isotope", "modele", "date_livraison", "decroissance", "palette", "date_depart"]
        entries.clear()

        for i, label in enumerate(labels):
            tk.Label(form_frame, text=label).grid(row=i, column=0, sticky="e")
            entry = tk.Entry(form_frame, width=40)
            entry.grid(row=i, column=1, padx=5, pady=2)
            if data:
                value = data[i] if i < len(data) else ""
                entry.insert(0, str(value) if value is not None else '')
            entries.append(entry)

        def valider():
            vals = [entries[c].get() for c in range(8)]  # Ne pas inclure PJ pour le moment
            if not vals[0] or not vals[1]:
                messagebox.showerror("Champs obligatoires", "Les champs 'n_venus' et 'lot' sont obligatoires.")
                return
            cursor = conn.cursor()
            if data is None:
                cursor.execute(
                    "INSERT INTO genes_uses (n_venus, lot, isotope, modele, date_livraison, decroissance, palette, date_depart) "
                    "VALUES (?,?,?,?,?,?,?,?)", vals)
            else:
                key = data[0]
                cursor.execute(
                    "UPDATE genes_uses SET lot=?, isotope=?, modele=?, date_livraison=?, decroissance=?, palette=?, date_depart=? "
                    "WHERE n_venus=?", vals[1:] + [key])
            conn.commit()
            refresh()

        tk.Button(form_frame, text="Valider", command=valider).grid(row=len(labels), column=0, columnspan=2, pady=10)

    def delete_selected():
        sel = listbox.curselection()
        if not sel:
            return
        idx = sel[0]
        row, _, _ = data_by_index[idx]
        key = row[0]
        cursor = conn.cursor()
        cursor.execute("DELETE FROM genes_uses WHERE n_venus=?", (key,))
        conn.commit()
        refresh()

    listbox.bind('<<ListboxSelect>>', show_details)
